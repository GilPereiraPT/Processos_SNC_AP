# app.py
import io
import pandas as pd
import streamlit as st
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

st.set_page_config(page_title="Importar TXT para Excel", layout="centered")

st.title("Importar TXT → Excel (formatado)")
st.caption("Versão Streamlit (PT-PT). Carrega o TXT com registos em largura fixa.")

uploaded = st.file_uploader("Escolhe o ficheiro TXT", type=["txt"])

def to_int_str(x: str) -> str:
    try:
        return str(int(x))
    except Exception:
        return x

def to_float(x: str):
    try:
        return float(x.replace(",", "."))
    except Exception:
        return None

def to_date(x: str):
    # x no formato DDMMAAAA (8 chars)
    if len(x) == 8 and x.isdigit():
        try:
            return datetime.strptime(x, "%d%m%Y")  # devolver datetime p/ formatar no Excel
        except Exception:
            return x
    return x

def parse_txt(content: str) -> pd.DataFrame:
    dados = []
    for raw in content.splitlines():
        linha = raw.rstrip("\n\r")
        # Precisamos de pelo menos 192 chars para apanhar a CC (183–192)
        if len(linha) >= 192:
            COD      = linha[0:3].strip()          # 1–3
            Entidade = linha[11:21].strip()        # 12–21
            NE       = linha[21:55].strip()        # 22–55
            DataStr  = linha[55:63].strip()        # 56–63
            Deb      = linha[63:113].strip()       # 64–113 (texto)
            Cred     = linha[113:155].strip()      # 114–155 (texto)
            Valor    = linha[155:181].strip()      # 156–181
            Sinal    = linha[181:182].strip()      # 182
            CC       = linha[182:192].strip()      # 183–192

            dados.append([
                COD,
                to_int_str(Entidade),
                NE,
                to_date(DataStr),
                Deb,
                Cred,
                to_float(Valor),
                Sinal,
                CC
            ])
    cols = ["COD", "Entidade", "NE", "Data", "Deb", "Cred", "Valor", "Sinal", "CC"]
    df = pd.DataFrame(dados, columns=cols)
    # Adiciona linha Total
    total = df["Valor"].sum(skipna=True) if not df.empty else 0.0
    df_total = df.copy()
    df_total.loc[len(df_total)] = ["", "", "", "", "", "Total", total, "", ""]
    return df, df_total

def format_excel_and_get_bytes(df_total: pd.DataFrame) -> bytes:
    # 1) Exportar para um BytesIO via pandas
    bio = io.BytesIO()
    df_total.to_excel(bio, index=False)
    bio.seek(0)

    # 2) Reabrir com openpyxl e aplicar formatação
    wb = load_workbook(bio)
    ws = wb.active

    # Cabeçalhos
    header_fill = PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Coluna Data (D) — formatar apenas células com datetime
    for cell in ws["D"][1:ws.max_row-1]:
        if hasattr(cell.value, "year"):  # é datetime
            cell.number_format = "DD/MM/YYYY"

    # Coluna Valor (G)
    for cell in ws["G"][1:ws.max_row]:
        if cell.value is not None and cell.value != "Total":
            cell.number_format = "#,##0.00"

    # Débito e Crédito → alinhar à esquerda (E, F)
    for col in ["E", "F"]:
        for cell in ws[col][1:ws.max_row-1]:
            cell.alignment = Alignment(horizontal="left")

    # Linha total (última)
    total_row = ws[ws.max_row]
    for cell in total_row:
        cell.font = Font(bold=True)
    thin = Side(style="thin")
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(ws.max_row, col_idx).border = Border(top=thin)

    # Ajustar larguras de coluna (simples)
    for column_cells in ws.columns:
        max_len = 0
        col_letter = column_cells[0].column_letter
        for c in column_cells:
            v = "" if c.value is None else str(c.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = max_len + 2

    # 3) Guardar novamente para bytes
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

if uploaded is not None:
    # tentar UTF-8 e fallback para cp1252/latin-1
    raw_bytes = uploaded.read()
    try:
        text = raw_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = raw_bytes.decode("cp1252", errors="replace")

    df, df_total = parse_txt(text)

    if df.empty:
        st.error("Nenhuma linha válida encontrada (verifica o layout/posições).")
    else:
        st.success(f"Importadas {len(df)} linhas de dados (sem contar com a linha Total).")
        st.subheader("Pré-visualização")
        st.dataframe(df_total, use_container_width=True)

        excel_bytes = format_excel_and_get_bytes(df_total)
        st.download_button(
            label="⬇️ Descarregar Excel formatado",
            data=excel_bytes,
            file_name=f"{uploaded.name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.info("Carrega um ficheiro TXT para começar.")

st.caption("Dica: se o teu TXT vier com outra codificação, a app tenta automaticamente cp1252 se UTF-8 falhar.")

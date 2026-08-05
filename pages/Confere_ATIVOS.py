from __future__ import annotations

import io
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from typing import BinaryIO, Iterable

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURAÇÃO DA PÁGINA
# ============================================================

st.set_page_config(
    page_title="Conferência Contabilidade × Ativos",
    page_icon="📊",
    layout="wide",
)

TOLERANCIA_PREDEFINIDA = 0.10


# ============================================================
# FUNÇÕES GERAIS
# ============================================================


def normalizar_texto(valor: object) -> str:
    texto = "" if valor is None else str(valor)
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", texto.strip().lower())


def normalizar_codigo(valor: object) -> str:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""

    texto = str(valor).strip()
    if texto.endswith(".0"):
        texto = texto[:-2]

    return re.sub(r"[^0-9A-Za-z]", "", texto)


def converter_montante(valor: object) -> float:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return 0.0

    if isinstance(valor, (int, float, Decimal)):
        return float(valor)

    texto = (
        str(valor)
        .strip()
        .replace("€", "")
        .replace("\u00a0", "")
        .replace(" ", "")
    )

    if texto in {"", "-", "—"}:
        return 0.0

    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")

    try:
        return float(Decimal(texto))
    except (InvalidOperation, ValueError):
        return 0.0


def formatar_euro(valor: float) -> str:
    return f"{valor:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")


def descodificar_csv(dados: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1252", "latin1"):
        try:
            texto = dados.decode(encoding)
            if ";" in texto:
                return texto
        except UnicodeDecodeError:
            continue

    return dados.decode("latin1", errors="replace")


def encontrar_coluna(colunas: Iterable[str], alternativas: Iterable[str]) -> str | None:
    mapa = {normalizar_texto(c): c for c in colunas}

    for alternativa in alternativas:
        chave = normalizar_texto(alternativa)
        if chave in mapa:
            return mapa[chave]

    return None


def e_conta_aft(codigo: str) -> bool:
    return bool(re.match(r"^43[1-7]", codigo))


def e_conta_ai(codigo: str) -> bool:
    return codigo.startswith("443")


def natureza_ativo(codigo: str) -> str:
    if e_conta_aft(codigo):
        return "Ativo fixo tangível"
    if e_conta_ai(codigo):
        return "Ativo intangível"
    return "Fora do âmbito"


def contas_finais(codigos: Iterable[str]) -> set[str]:
    lista = sorted({str(c) for c in codigos if str(c)})
    return {
        codigo
        for codigo in lista
        if not any(outro.startswith(codigo) and outro != codigo for outro in lista)
    }


# ============================================================
# LEITURA DO BALANCETE SICC
# ============================================================


def carregar_sicc(ficheiro: BinaryIO) -> pd.DataFrame:
    dados = ficheiro.read()
    texto = descodificar_csv(dados)
    linhas = texto.splitlines()

    indice_cabecalho = next(
        (
            i
            for i, linha in enumerate(linhas)
            if normalizar_texto(linha).startswith("conta;designacao da conta")
        ),
        None,
    )

    if indice_cabecalho is None:
        raise ValueError("Não foi encontrado o cabeçalho do balancete SICC.")

    df = pd.read_csv(
        io.StringIO("\n".join(linhas[indice_cabecalho:])),
        sep=";",
        dtype=str,
    )
    df = df.loc[:, ~df.columns.astype(str).str.match(r"^Unnamed")]
    df.columns = [normalizar_texto(c) for c in df.columns]

    conta_col = encontrar_coluna(df.columns, ["Conta"])
    descricao_col = encontrar_coluna(df.columns, ["Designação da conta", "Designacao da conta"])
    valor_debito_col = encontrar_coluna(
        df.columns,
        ["Valor a débito", "Valor a debito", "Débito", "Debito"],
    )
    valor_credito_col = encontrar_coluna(
        df.columns,
        ["Valor a crédito", "Valor a credito", "Crédito", "Credito"],
    )
    saldo_debito_col = encontrar_coluna(df.columns, ["Saldo a débito", "Saldo a debito"])
    saldo_credito_col = encontrar_coluna(df.columns, ["Saldo a crédito", "Saldo a credito"])
    acumulado_debito_col = encontrar_coluna(
        df.columns,
        ["Valor acumulado a débito", "Valor acumulado a debito"],
    )
    acumulado_credito_col = encontrar_coluna(
        df.columns,
        ["Valor acumulado a crédito", "Valor acumulado a credito"],
    )

    obrigatorias = {
        "Conta": conta_col,
        "Designação da conta": descricao_col,
        "Saldo a débito": saldo_debito_col,
        "Saldo a crédito": saldo_credito_col,
    }
    em_falta = [nome for nome, coluna in obrigatorias.items() if coluna is None]

    if em_falta:
        raise ValueError(f"Faltam colunas obrigatórias no SICC: {', '.join(em_falta)}")

    # Há versões do SICC em que o movimento do período surge como
    # "Valor a débito/crédito" e outras em que apenas aparece o acumulado.
    # O período só é calculado quando as colunas próprias existem.
    out = pd.DataFrame(
        {
            "conta": df[conta_col].map(normalizar_codigo),
            "descricao": df[descricao_col].fillna(""),
            "valor_debito_periodo": (
                df[valor_debito_col].map(converter_montante)
                if valor_debito_col
                else 0.0
            ),
            "valor_credito_periodo": (
                df[valor_credito_col].map(converter_montante)
                if valor_credito_col
                else 0.0
            ),
            "saldo_debito": df[saldo_debito_col].map(converter_montante),
            "saldo_credito": df[saldo_credito_col].map(converter_montante),
            "valor_acumulado_debito": (
                df[acumulado_debito_col].map(converter_montante)
                if acumulado_debito_col
                else df[saldo_debito_col].map(converter_montante)
            ),
            "valor_acumulado_credito": (
                df[acumulado_credito_col].map(converter_montante)
                if acumulado_credito_col
                else df[saldo_credito_col].map(converter_montante)
            ),
        }
    )

    out["movimento_periodo_liquido"] = (
        out["valor_debito_periodo"] - out["valor_credito_periodo"]
    )
    out["saldo_liquido_devedor"] = out["saldo_debito"] - out["saldo_credito"]
    out["saldo_liquido_credor"] = out["saldo_credito"] - out["saldo_debito"]

    out["tem_colunas_periodo"] = bool(valor_debito_col and valor_credito_col)
    out = out[out["conta"] != ""].reset_index(drop=True)

    if out.empty:
        raise ValueError("O balancete SICC não contém contas válidas.")

    return out


# ============================================================
# LEITURA DO BALANCETE PRIMAVERA
# ============================================================


def encontrar_cabecalho_primavera(raw: pd.DataFrame) -> int:
    for idx in range(min(40, len(raw))):
        valores = {normalizar_texto(v) for v in raw.iloc[idx].tolist() if pd.notna(v)}
        if {"codigo", "descricao", "valor contabilistico"}.issubset(valores):
            return idx

    raise ValueError("Não foi encontrado o cabeçalho do balancete Primavera.")


def carregar_primavera(ficheiro: BinaryIO) -> tuple[pd.DataFrame, pd.DataFrame]:
    dados = ficheiro.read()
    raw = pd.read_excel(io.BytesIO(dados), header=None, engine="openpyxl")
    cabecalho_idx = encontrar_cabecalho_primavera(raw)

    cabecalho = [normalizar_texto(v) for v in raw.iloc[cabecalho_idx].tolist()]
    df = raw.iloc[cabecalho_idx + 1 :].copy()
    df.columns = cabecalho

    colunas = list(df.columns)

    try:
        pos_codigo = colunas.index("codigo")
        pos_descricao = colunas.index("descricao")
        pos_data = colunas.index("data utilizacao")
        pos_valor = colunas.index("valor contabilistico")
        pos_residual = colunas.index("valor residual")
        pos_taxa = colunas.index("taxa")
        pos_quantia = colunas.index("quantia escriturada")
    except ValueError as exc:
        raise ValueError(f"Estrutura inesperada no ficheiro Primavera: {exc}") from exc

    # Estrutura conhecida do relatório Primavera:
    # 7 Período depreciação; 8 Exercício; 9 Acumulada;
    # 10 Período imparidade; 11 Exercício; 12 Acumulada.
    if len(colunas) < 14:
        raise ValueError("O balancete Primavera não contém todas as colunas esperadas.")

    pos_dep_periodo = 7
    pos_dep_exercicio = 8
    pos_dep_acumulada = 9
    pos_imp_periodo = 10
    pos_imp_exercicio = 11
    pos_imp_acumulada = 12

    linhas = pd.DataFrame(
        {
            "codigo_original": df.iloc[:, pos_codigo],
            "codigo": df.iloc[:, pos_codigo].map(normalizar_codigo),
            "descricao": df.iloc[:, pos_descricao].fillna(""),
            "data_utilizacao": pd.to_datetime(df.iloc[:, pos_data], errors="coerce"),
            "valor_contabilistico": df.iloc[:, pos_valor].map(converter_montante),
            "valor_residual": df.iloc[:, pos_residual].map(converter_montante),
            "taxa": df.iloc[:, pos_taxa].map(converter_montante),
            "depreciacao_periodo": df.iloc[:, pos_dep_periodo].map(converter_montante),
            "depreciacao_exercicio": df.iloc[:, pos_dep_exercicio].map(converter_montante),
            "depreciacao_acumulada": df.iloc[:, pos_dep_acumulada].map(converter_montante),
            "imparidade_periodo": df.iloc[:, pos_imp_periodo].map(converter_montante),
            "imparidade_exercicio": df.iloc[:, pos_imp_exercicio].map(converter_montante),
            "imparidade_acumulada": df.iloc[:, pos_imp_acumulada].map(converter_montante),
            "quantia_escriturada": df.iloc[:, pos_quantia].map(converter_montante),
        }
    )

    linhas["e_conta"] = linhas["codigo_original"].map(
        lambda v: bool(
            v is not None
            and not (isinstance(v, float) and pd.isna(v))
            and not str(v).startswith(" ")
        )
    )

    linhas = linhas[linhas["codigo"] != ""].reset_index(drop=True)

    # Atribui cada ficha individual à conta contabilística imediatamente anterior.
    conta_corrente = ""
    contas_ficha: list[str] = []

    for _, linha in linhas.iterrows():
        codigo = linha["codigo"]
        if linha["e_conta"]:
            conta_corrente = codigo
            contas_ficha.append(codigo)
        else:
            contas_ficha.append(conta_corrente)

    linhas["conta_ativo"] = contas_ficha
    linhas["natureza"] = linhas["conta_ativo"].map(natureza_ativo)

    contas = linhas[
        linhas["e_conta"]
        & (
            linhas["codigo"].map(e_conta_aft)
            | linhas["codigo"].map(e_conta_ai)
            | linhas["codigo"].isin(["43", "44"])
        )
    ].copy()

    fichas = linhas[
        ~linhas["e_conta"]
        & (linhas["conta_ativo"].map(e_conta_aft) | linhas["conta_ativo"].map(e_conta_ai))
    ].copy()

    return contas.reset_index(drop=True), fichas.reset_index(drop=True)


# ============================================================
# MAPEAMENTOS CONTABILÍSTICOS
# ============================================================


# Mapeamentos específicos têm prioridade sobre a inferência por raiz.
# Uma conta de depreciação pode agregar várias contas de aquisição.
MAPA_GASTOS_AFT: dict[str, list[str]] = {
    "6422": ["432"],
    "642331": ["43331"],
    "642332": ["43332"],
    "642333": ["43333"],
    "642334": ["43334"],
    "642335": ["43335"],
    "642339": ["43339"],
    "64235": ["4335"],
    "64239": ["4339"],
    "6424": ["434"],
    "642511": ["43511"],
    # Regra específica: estas três contas acumulam conjuntamente em 64259.
    "64259": ["4352", "4353", "4359"],
    "6427": ["437"],
}

MAPA_ACUMULADAS_AFT: dict[str, list[str]] = {
    "4382": ["432"],
    "438331": ["43331"],
    "438332": ["43332"],
    "438333": ["43333"],
    "438334": ["43334"],
    "438335": ["43335"],
    "438339": ["43339"],
    "43835": ["4335"],
    "43839": ["4339"],
    "4384": ["434"],
    "438511": ["43511"],
    # Regra específica: 4352, 4353 e 4359 acumulam conjuntamente em 43859.
    "43859": ["4352", "4353", "4359"],
    "4387": ["437"],
}

MAPA_GASTOS_AI: dict[str, list[str]] = {
    "6433": ["443"],
}

MAPA_ACUMULADAS_AI: dict[str, list[str]] = {
    "4483": ["443"],
}


def raiz_base_por_gasto(conta_gasto: str) -> str:
    """
    Converte a conta de gasto na raiz contabilística do ativo.

    Exemplos:
    6422   -> 432   (abrange 4321, 4324, ...)
    642331 -> 43331
    6424   -> 434
    6433   -> 443
    """
    conta_gasto = normalizar_codigo(conta_gasto)
    if conta_gasto.startswith("642"):
        return "43" + conta_gasto[3:]
    if conta_gasto.startswith("643"):
        return "443"
    return ""


def raiz_base_por_acumulada(conta_acumulada: str) -> str:
    """
    Converte a conta de depreciação/amortização acumulada na raiz do ativo.

    Exemplos:
    4382   -> 432   (abrange 4321, 4324, ...)
    438331 -> 43331
    4384   -> 434
    4483   -> 443
    """
    conta_acumulada = normalizar_codigo(conta_acumulada)
    if conta_acumulada.startswith("438"):
        return "43" + conta_acumulada[3:]
    if conta_acumulada.startswith("4483"):
        return "443"
    return ""


def resolver_raiz_primavera(
    contas_primavera: pd.DataFrame,
    raiz_proposta: str,
) -> str:
    """
    Resolve a raiz efetivamente existente no Primavera.

    A conta de amortização pode ter menos detalhe do que as contas de aquisição.
    Nesse caso mantém-se a raiz mais específica que possua contas descendentes
    no Primavera. Se a raiz proposta não existir, recua progressivamente na
    hierarquia, sem sair do grupo 431-437 ou 443.
    """
    raiz = normalizar_codigo(raiz_proposta)
    if not raiz:
        return ""

    codigos = contas_primavera["codigo"].astype(str)

    def existe(prefixo: str) -> bool:
        return bool(codigos.str.startswith(prefixo).any())

    if existe(raiz):
        return raiz

    minimo = 3 if raiz.startswith("43") else len(raiz)
    while len(raiz) > minimo:
        raiz = raiz[:-1]
        if (e_conta_aft(raiz) or e_conta_ai(raiz)) and existe(raiz):
            return raiz

    return raiz_proposta


def contas_primavera_abrangidas(
    contas_primavera: pd.DataFrame,
    raiz: str,
) -> list[str]:
    """Lista apenas as contas finais do Primavera abrangidas pela raiz."""
    grupo = contas_primavera[
        contas_primavera["codigo"].astype(str).str.startswith(raiz)
    ].copy()
    if grupo.empty:
        return []
    finais = contas_finais(grupo["codigo"])
    return sorted(finais)


def somar_primavera_por_raiz(
    contas_primavera: pd.DataFrame,
    raiz: str,
    coluna: str,
) -> float:
    """
    Soma todas as contas finais do Primavera pertencentes à raiz.

    Para reconciliação de amortizações não usa automaticamente a linha-mãe,
    porque a conta SICC pode agregar várias contas de aquisição do Primavera.
    Assim, 4382/6422 compara com a soma das contas finais 432..., sem duplicar
    linhas agregadoras e subcontas.
    """
    grupo = contas_primavera[
        contas_primavera["codigo"].astype(str).str.startswith(raiz)
    ].copy()
    if grupo.empty:
        return 0.0

    finais = contas_finais(grupo["codigo"])
    if finais:
        return float(grupo[grupo["codigo"].isin(finais)][coluna].sum())

    exata = grupo[grupo["codigo"] == raiz]
    return float(exata[coluna].sum()) if not exata.empty else 0.0


def resolver_raizes_mapeadas(
    contas_primavera: pd.DataFrame,
    raizes_propostas: list[str],
) -> list[str]:
    """Resolve várias raízes e remove duplicados, preservando a ordem."""
    resultado: list[str] = []
    for raiz in raizes_propostas:
        resolvida = resolver_raiz_primavera(contas_primavera, raiz)
        if resolvida and resolvida not in resultado:
            resultado.append(resolvida)
    return resultado


def contas_primavera_abrangidas_por_raizes(
    contas_primavera: pd.DataFrame,
    raizes: list[str],
) -> list[str]:
    """Lista contas finais abrangidas por várias raízes, sem duplicação."""
    contas: set[str] = set()
    for raiz in raizes:
        contas.update(contas_primavera_abrangidas(contas_primavera, raiz))
    return sorted(contas)


def somar_primavera_por_raizes(
    contas_primavera: pd.DataFrame,
    raizes: list[str],
    coluna: str,
) -> float:
    """Soma várias raízes contabilísticas, evitando contar contas duas vezes."""
    codigos_finais = contas_primavera_abrangidas_por_raizes(
        contas_primavera, raizes
    )
    if codigos_finais:
        return float(
            contas_primavera.loc[
                contas_primavera["codigo"].isin(codigos_finais), coluna
            ].sum()
        )

    # Fallback para linhas exatas quando não existem descendentes finais.
    exatas = contas_primavera[contas_primavera["codigo"].isin(raizes)]
    return float(exatas[coluna].sum()) if not exatas.empty else 0.0


def descricao_primavera_por_raizes(
    contas_primavera: pd.DataFrame,
    raizes: list[str],
) -> str:
    descricoes: list[str] = []
    for raiz in raizes:
        descricao = descricao_primavera_por_prefixo(contas_primavera, raiz)
        if descricao and descricao not in descricoes:
            descricoes.append(descricao)
    return " | ".join(descricoes)


def somar_primavera_por_prefixo(
    contas_primavera: pd.DataFrame,
    prefixo: str,
    coluna: str,
) -> float:
    """
    Para o valor contabilístico, usa a conta exata quando existe; caso contrário,
    soma apenas contas finais descendentes.
    """
    exata = contas_primavera[contas_primavera["codigo"] == prefixo]
    if not exata.empty:
        return float(exata.iloc[0][coluna])

    descendentes = contas_primavera[
        contas_primavera["codigo"].astype(str).str.startswith(prefixo)
    ].copy()
    if descendentes.empty:
        return 0.0

    finais = contas_finais(descendentes["codigo"])
    return float(descendentes[descendentes["codigo"].isin(finais)][coluna].sum())


def somar_sicc_grupo_sem_duplicacao(
    sicc: pd.DataFrame,
    prefixo: str,
    coluna: str,
) -> float:
    """
    Obtém o total de um grupo contabilístico sem duplicar contas-mãe e subcontas.

    Regra:
    1. Se existir a conta agregadora exata (por exemplo, 438), usa essa linha.
    2. Se não existir, soma apenas as contas finais descendentes.
    3. Se houver linhas repetidas para a mesma conta, agrega-as primeiro.
    """
    grupo = sicc[sicc["conta"].str.startswith(prefixo)].copy()
    if grupo.empty:
        return 0.0

    grupo = (
        grupo.groupby("conta", as_index=False)[coluna]
        .sum()
    )

    exata = grupo[grupo["conta"] == prefixo]
    if not exata.empty:
        return float(exata[coluna].sum())

    finais = contas_finais(grupo["conta"])
    return float(grupo[grupo["conta"].isin(finais)][coluna].sum())


def somar_primavera_por_natureza(
    contas_primavera: pd.DataFrame,
    natureza: str,
    coluna: str,
) -> float:
    """Soma o Primavera por natureza sem duplicar níveis hierárquicos."""
    if natureza == "Ativo fixo tangível":
        return float(sum(
            somar_primavera_por_prefixo(contas_primavera, prefixo, coluna)
            for prefixo in ("431", "432", "433", "434", "435", "436", "437")
        ))

    if natureza == "Ativo intangível":
        return float(somar_primavera_por_prefixo(contas_primavera, "443", coluna))

    return 0.0

def descricao_primavera_por_prefixo(contas_primavera: pd.DataFrame, prefixo: str) -> str:
    exata = contas_primavera[contas_primavera["codigo"] == prefixo]
    if not exata.empty:
        return str(exata.iloc[0]["descricao"])
    return ""


# ============================================================
# RECONCILIAÇÃO CONTABILÍSTICA
# ============================================================


def reconciliar_contas(
    sicc: pd.DataFrame,
    contas_primavera: pd.DataFrame,
    tolerancia: float,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    detalhes: list[dict] = []

    # --------------------------------------------------------
    # 1. Valor contabilístico dos ativos
    # AFT: 431...437
    # AI:  443...
    # SICC: saldo a débito - saldo a crédito
    # --------------------------------------------------------
    contas_valor = sicc[
        sicc["conta"].map(e_conta_aft) | sicc["conta"].map(e_conta_ai)
    ].copy()

    finais_valor = contas_finais(contas_valor["conta"])
    contas_valor = contas_valor[contas_valor["conta"].isin(finais_valor)]

    for _, linha in contas_valor.iterrows():
        conta = linha["conta"]
        valor_sicc = float(linha["saldo_liquido_devedor"])
        valor_primavera = somar_primavera_por_prefixo(
            contas_primavera,
            conta,
            "valor_contabilistico",
        )
        diferenca = valor_sicc - valor_primavera

        detalhes.append(
            {
                "Componente": "Valor contabilístico",
                "Natureza": natureza_ativo(conta),
                "Conta SICC": conta,
                "Conta Primavera": conta,
                "Contas Primavera abrangidas": conta,
                "Descrição SICC": linha["descricao"],
                "Descrição Primavera": descricao_primavera_por_prefixo(contas_primavera, conta),
                "Cálculo SICC": "Saldo a débito - Saldo a crédito",
                "SICC": valor_sicc,
                "Primavera": valor_primavera,
                "Diferença SICC - Primavera": diferenca,
                "Estado": "OK" if abs(diferenca) <= tolerancia else "Divergência",
            }
        )

    # --------------------------------------------------------
    # 2. Depreciações/amortizações do período e exercício
    # AFT: 642...
    # AI:  643...
    # Período: Valor a débito - Valor a crédito
    # Exercício: Saldo a débito - Saldo a crédito
    # --------------------------------------------------------
    contas_gasto = sicc[sicc["conta"].str.startswith(("642", "643"))].copy()
    finais_gasto = contas_finais(contas_gasto["conta"])
    contas_gasto = contas_gasto[contas_gasto["conta"].isin(finais_gasto)]

    for _, linha in contas_gasto.iterrows():
        conta_sicc = linha["conta"]
        mapa_gastos = {**MAPA_GASTOS_AFT, **MAPA_GASTOS_AI}
        raizes_propostas = mapa_gastos.get(
            conta_sicc,
            [raiz_base_por_gasto(conta_sicc)],
        )
        raizes_primavera = resolver_raizes_mapeadas(
            contas_primavera, raizes_propostas
        )
        conta_primavera = " + ".join(raizes_primavera)
        natureza = natureza_ativo(raizes_primavera[0]) if raizes_primavera else ""

        periodo_sicc = float(linha["movimento_periodo_liquido"])
        periodo_primavera = somar_primavera_por_raizes(
            contas_primavera,
            raizes_primavera,
            "depreciacao_periodo",
        )
        dif_periodo = periodo_sicc - periodo_primavera

        detalhes.append(
            {
                "Componente": "Depreciação/amortização do período",
                "Natureza": natureza,
                "Conta SICC": conta_sicc,
                "Conta Primavera": conta_primavera,
                "Contas Primavera abrangidas": ", ".join(
                    contas_primavera_abrangidas_por_raizes(
                        contas_primavera, raizes_primavera
                    )
                ),
                "Descrição SICC": linha["descricao"],
                "Descrição Primavera": descricao_primavera_por_raizes(
                    contas_primavera, raizes_primavera
                ),
                "Cálculo SICC": "Valor a débito - Valor a crédito",
                "SICC": periodo_sicc,
                "Primavera": periodo_primavera,
                "Diferença SICC - Primavera": dif_periodo,
                "Estado": "OK" if abs(dif_periodo) <= tolerancia else "Divergência",
            }
        )

        exercicio_sicc = float(linha["saldo_liquido_devedor"])
        exercicio_primavera = somar_primavera_por_raizes(
            contas_primavera,
            raizes_primavera,
            "depreciacao_exercicio",
        )
        dif_exercicio = exercicio_sicc - exercicio_primavera

        detalhes.append(
            {
                "Componente": "Depreciação/amortização do exercício",
                "Natureza": natureza,
                "Conta SICC": conta_sicc,
                "Conta Primavera": conta_primavera,
                "Contas Primavera abrangidas": ", ".join(
                    contas_primavera_abrangidas_por_raizes(
                        contas_primavera, raizes_primavera
                    )
                ),
                "Descrição SICC": linha["descricao"],
                "Descrição Primavera": descricao_primavera_por_raizes(
                    contas_primavera, raizes_primavera
                ),
                "Cálculo SICC": "Saldo a débito - Saldo a crédito",
                "SICC": exercicio_sicc,
                "Primavera": exercicio_primavera,
                "Diferença SICC - Primavera": dif_exercicio,
                "Estado": "OK" if abs(dif_exercicio) <= tolerancia else "Divergência",
            }
        )

    # --------------------------------------------------------
    # 3. Depreciações/amortizações acumuladas
    # AFT: 438...
    # AI:  4483...
    # SICC: saldo a crédito - saldo a débito
    # --------------------------------------------------------
    contas_acumuladas = sicc[
        sicc["conta"].str.startswith("438")
        | sicc["conta"].str.startswith("4483")
    ].copy()
    finais_acumuladas = contas_finais(contas_acumuladas["conta"])
    contas_acumuladas = contas_acumuladas[
        contas_acumuladas["conta"].isin(finais_acumuladas)
    ]

    for _, linha in contas_acumuladas.iterrows():
        conta_sicc = linha["conta"]
        mapa_acumuladas = {**MAPA_ACUMULADAS_AFT, **MAPA_ACUMULADAS_AI}
        raizes_propostas = mapa_acumuladas.get(
            conta_sicc,
            [raiz_base_por_acumulada(conta_sicc)],
        )
        raizes_primavera = resolver_raizes_mapeadas(
            contas_primavera, raizes_propostas
        )
        conta_primavera = " + ".join(raizes_primavera)
        valor_sicc = float(linha["saldo_liquido_credor"])
        valor_primavera = somar_primavera_por_raizes(
            contas_primavera,
            raizes_primavera,
            "depreciacao_acumulada",
        )
        diferenca = valor_sicc - valor_primavera

        detalhes.append(
            {
                "Componente": "Depreciação/amortização acumulada",
                "Natureza": natureza_ativo(raizes_primavera[0]) if raizes_primavera else "",
                "Conta SICC": conta_sicc,
                "Conta Primavera": conta_primavera,
                "Contas Primavera abrangidas": ", ".join(
                    contas_primavera_abrangidas_por_raizes(
                        contas_primavera, raizes_primavera
                    )
                ),
                "Descrição SICC": linha["descricao"],
                "Descrição Primavera": descricao_primavera_por_raizes(
                    contas_primavera, raizes_primavera
                ),
                "Cálculo SICC": "Saldo a crédito - Saldo a débito",
                "SICC": valor_sicc,
                "Primavera": valor_primavera,
                "Diferença SICC - Primavera": diferenca,
                "Estado": "OK" if abs(diferenca) <= tolerancia else "Divergência",
            }
        )

    detalhe = pd.DataFrame(detalhes)

    if detalhe.empty:
        detalhe = pd.DataFrame(
            columns=[
                "Componente",
                "Natureza",
                "Conta SICC",
                "Conta Primavera",
                "Descrição SICC",
                "Descrição Primavera",
                "Cálculo SICC",
                "SICC",
                "Primavera",
                "Diferença SICC - Primavera",
                "Estado",
            ]
        )

    # Resumo por componente e natureza.
    if detalhe.empty:
        resumo = pd.DataFrame(
            columns=[
                "Componente",
                "Natureza",
                "SICC",
                "Primavera",
                "Diferença SICC - Primavera",
                "Estado",
            ]
        )
    else:
        resumo = (
            detalhe.groupby(["Componente", "Natureza"], as_index=False)[["SICC", "Primavera"]]
            .sum()
        )
        # Corrige os totais das depreciações/amortizações acumuladas.
        # O detalhe é apresentado por contas finais, mas o resumo deve usar a
        # conta agregadora quando ela existe (438 para AFT e 4483 para AI).
        # Isto impede a dupla contagem de contas-mãe e subcontas.
        ajustes_acumuladas = [
            (
                "Ativo fixo tangível",
                "438",
                somar_sicc_grupo_sem_duplicacao(
                    sicc, "438", "saldo_liquido_credor"
                ),
            ),
            (
                "Ativo intangível",
                "4483",
                somar_sicc_grupo_sem_duplicacao(
                    sicc, "4483", "saldo_liquido_credor"
                ),
            ),
        ]

        for natureza, _prefixo_sicc, total_sicc in ajustes_acumuladas:
            mascara = (
                (resumo["Componente"] == "Depreciação/amortização acumulada")
                & (resumo["Natureza"] == natureza)
            )
            if mascara.any():
                resumo.loc[mascara, "SICC"] = total_sicc
                resumo.loc[mascara, "Primavera"] = somar_primavera_por_natureza(
                    contas_primavera,
                    natureza,
                    "depreciacao_acumulada",
                )

        resumo["Diferença SICC - Primavera"] = resumo["SICC"] - resumo["Primavera"]
        resumo["Estado"] = resumo["Diferença SICC - Primavera"].abs().map(
            lambda v: "OK" if v <= tolerancia else "Divergência"
        )

    cobertura = pd.DataFrame(
        [
            {
                "Grupo de contas": "431 a 437",
                "Finalidade": "Valor contabilístico dos AFT",
                "Disponível no SICC": bool(sicc["conta"].map(e_conta_aft).any()),
            },
            {
                "Grupo de contas": "443",
                "Finalidade": "Valor contabilístico dos ativos intangíveis",
                "Disponível no SICC": bool(sicc["conta"].map(e_conta_ai).any()),
            },
            {
                "Grupo de contas": "642",
                "Finalidade": "Depreciações dos AFT — período e exercício",
                "Disponível no SICC": bool(sicc["conta"].str.startswith("642").any()),
            },
            {
                "Grupo de contas": "643",
                "Finalidade": "Amortizações dos intangíveis — período e exercício",
                "Disponível no SICC": bool(sicc["conta"].str.startswith("643").any()),
            },
            {
                "Grupo de contas": "438",
                "Finalidade": "Depreciações acumuladas dos AFT",
                "Disponível no SICC": bool(sicc["conta"].str.startswith("438").any()),
            },
            {
                "Grupo de contas": "4483",
                "Finalidade": "Amortizações acumuladas dos ativos intangíveis",
                "Disponível no SICC": bool(sicc["conta"].str.startswith("4483").any()),
            },
        ]
    )

    return resumo, detalhe, cobertura


# ============================================================
# CONTROLO DOS BENS SEM DEPRECIAÇÃO/AMORTIZAÇÃO
# ============================================================


def controlar_fichas(
    fichas: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if fichas.empty:
        resumo = pd.DataFrame(
            [
                {
                    "Indicador": "Fichas individuais analisadas",
                    "Quantidade": 0,
                }
            ]
        )
        return resumo, pd.DataFrame()

    controlo = fichas.copy()
    controlo["descricao_normalizada"] = controlo["descricao"].map(normalizar_texto)

    controlo["terreno"] = (
        controlo["conta_ativo"].str.startswith("431")
        | controlo["descricao_normalizada"].str.contains(r"\bterreno\b", regex=True)
    )

    # Regra de controlo:
    # 1. Os terrenos não depreciam/amortizam.
    # 2. Nos restantes bens, se o valor de aquisição menos as depreciações/
    #    amortizações acumuladas for positivo, tem de existir um valor positivo
    #    de depreciação/amortização no período.
    #
    # A tolerância usada na reconciliação contabilística não é aplicada aqui:
    # para este controlo, qualquer valor remanescente superior a zero exige que
    # o valor do período seja também estritamente superior a zero.
    controlo["valor_aquisicao_menos_acumuladas"] = (
        controlo["valor_contabilistico"]
        - controlo["depreciacao_acumulada"]
    )
    controlo["tem_valor_por_amortizar"] = (
        controlo["valor_aquisicao_menos_acumuladas"] > 0
    )
    controlo["deve_amortizar_no_periodo"] = (
        ~controlo["terreno"]
        & controlo["tem_valor_por_amortizar"]
    )
    controlo["erro_sem_amortizacao_periodo"] = (
        controlo["deve_amortizar_no_periodo"]
        & (controlo["depreciacao_periodo"] <= 0)
    )

    controlo["motivo"] = ""
    controlo.loc[
        controlo["erro_sem_amortizacao_periodo"],
        "motivo",
    ] = (
        "Valor de aquisição menos amortizações/depreciações acumuladas é "
        "positivo, mas a amortização/depreciação do período não é superior a zero"
    )
    problemas = controlo[controlo["erro_sem_amortizacao_periodo"]].copy()

    colunas_problemas = [
        "codigo",
        "conta_ativo",
        "natureza",
        "descricao",
        "data_utilizacao",
        "taxa",
        "valor_contabilistico",
        "depreciacao_acumulada",
        "valor_aquisicao_menos_acumuladas",
        "depreciacao_periodo",
        "depreciacao_exercicio",
        "motivo",
    ]
    problemas = problemas[colunas_problemas].rename(
        columns={
            "codigo": "Ficha",
            "conta_ativo": "Conta do ativo",
            "natureza": "Natureza",
            "descricao": "Descrição",
            "data_utilizacao": "Data de utilização",
            "taxa": "Taxa",
            "valor_contabilistico": "Valor de aquisição",
            "depreciacao_acumulada": "Acumulada",
            "valor_aquisicao_menos_acumuladas": "Valor de aquisição - acumulada",
            "depreciacao_periodo": "Período",
            "depreciacao_exercicio": "Exercício",
            "motivo": "Motivo",
        }
    )

    resumo = pd.DataFrame(
        [
            {"Indicador": "Fichas individuais analisadas", "Quantidade": int(len(controlo))},
            {"Indicador": "Terrenos excluídos", "Quantidade": int(controlo["terreno"].sum())},
            {
                "Indicador": "Bens não terrenos com valor por amortizar",
                "Quantidade": int(controlo["deve_amortizar_no_periodo"].sum()),
            },
            {
                "Indicador": "Bens sem valor por amortizar",
                "Quantidade": int((~controlo["tem_valor_por_amortizar"]).sum()),
            },
            {
                "Indicador": "Erros: bens que deviam amortizar sem valor positivo no período",
                "Quantidade": int(len(problemas)),
            },
        ]
    )

    return resumo, problemas


# ============================================================
# GERAÇÃO SEGURA DO EXCEL
# ============================================================


def gerar_excel(
    resumo: pd.DataFrame,
    detalhe: pd.DataFrame,
    cobertura: pd.DataFrame,
    resumo_controlo: pd.DataFrame,
    problemas: pd.DataFrame,
    sicc: pd.DataFrame,
    contas_primavera: pd.DataFrame,
    fichas_primavera: pd.DataFrame,
) -> bytes:
    output = io.BytesIO()

    conjuntos = [
        ("Resumo", resumo),
        ("Contas divergentes", detalhe),
        ("Cobertura", cobertura),
        ("Resumo controlo", resumo_controlo),
        ("Bens a verificar", problemas),
        ("SICC normalizado", sicc),
        ("Primavera contas", contas_primavera),
        ("Primavera fichas", fichas_primavera),
    ]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        folhas_escritas = 0

        for nome_folha, dados in conjuntos:
            if dados is None:
                df = pd.DataFrame()
            elif isinstance(dados, pd.DataFrame):
                df = dados.copy()
            else:
                try:
                    df = pd.DataFrame(dados)
                except Exception:
                    df = pd.DataFrame(
                        {"Aviso": [f"Não foi possível converter o conteúdo de {nome_folha}."]}
                    )

            # A folha Resumo é sempre criada, ainda que não existam resultados.
            if nome_folha == "Resumo" and df.empty and len(df.columns) == 0:
                df = pd.DataFrame(
                    {
                        "Estado": ["Sem resultados"],
                        "Observação": [
                            "Os ficheiros foram processados, mas não foram encontradas contas conciliáveis."
                        ],
                    }
                )

            if nome_folha != "Resumo" and df.empty and len(df.columns) == 0:
                continue

            df.to_excel(writer, sheet_name=nome_folha[:31], index=False)
            folhas_escritas += 1

        # Proteção final contra workbook sem folhas visíveis.
        if folhas_escritas == 0:
            pd.DataFrame(
                {
                    "Estado": ["Relatório sem dados"],
                    "Observação": ["Não foi possível gerar folhas com os dados recebidos."],
                }
            ).to_excel(writer, sheet_name="Diagnóstico", index=False)

        for ws in writer.book.worksheets:
            ws.sheet_state = "visible"
            ws.freeze_panes = "A2"

            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions

            for coluna_idx in range(1, ws.max_column + 1):
                valores = [
                    str(ws.cell(row=linha_idx, column=coluna_idx).value or "")
                    for linha_idx in range(1, ws.max_row + 1)
                ]
                largura = min(max((len(v) for v in valores), default=0) + 2, 45)
                ws.column_dimensions[get_column_letter(coluna_idx)].width = max(largura, 10)

        writer.book.active = 0

    output.seek(0)
    return output.getvalue()


# ============================================================
# FORMATAÇÃO STREAMLIT
# ============================================================


def estilizar_tabela(df: pd.DataFrame, colunas_monetarias: list[str]):
    formatos = {col: "{:,.2f} €" for col in colunas_monetarias if col in df.columns}
    styler = df.style.format(formatos)

    if "Estado" in df.columns:
        styler = styler.map(
            lambda valor: (
                "background-color: #ffe5e5"
                if valor == "Divergência"
                else "background-color: #e7f6e7"
                if valor == "OK"
                else ""
            ),
            subset=["Estado"],
        )

    return styler


# ============================================================
# INTERFACE
# ============================================================

st.title("Conferência do Balancete Contabilístico com o Registo de Ativos")
st.caption(
    "Reconcilia AFT e ativos intangíveis entre o SICC e o Primavera e identifica bens potencialmente sem depreciação ou amortização."
)

with st.sidebar:
    st.header("Parâmetros")
    tolerancia = st.number_input(
        "Tolerância (€)",
        min_value=0.0,
        value=TOLERANCIA_PREDEFINIDA,
        step=0.01,
        format="%.2f",
    )
    apenas_divergencias = st.checkbox("Mostrar apenas divergências", value=True)

coluna_1, coluna_2 = st.columns(2)

with coluna_1:
    ficheiro_sicc = st.file_uploader(
        "Balancete da contabilidade — SICC (CSV)",
        type=["csv"],
    )

with coluna_2:
    ficheiro_primavera = st.file_uploader(
        "Balancete do registo de ativos — Primavera (XLSX)",
        type=["xlsx"],
    )

if ficheiro_sicc and ficheiro_primavera:
    try:
        sicc = carregar_sicc(ficheiro_sicc)
        contas_primavera, fichas_primavera = carregar_primavera(ficheiro_primavera)

        resumo, detalhe, cobertura = reconciliar_contas(
            sicc,
            contas_primavera,
            tolerancia,
        )

        resumo_controlo, problemas = controlar_fichas(
            fichas_primavera,
        )

        separador_1, separador_2, separador_3, separador_4 = st.tabs(
            [
                "Resumo",
                "Divergências por conta",
                "Bens sem amortização",
                "Cobertura e lógica",
            ]
        )

        with separador_1:
            st.subheader("Resumo da reconciliação")

            if resumo.empty:
                st.warning("Não foram encontradas contas conciliáveis nos ficheiros carregados.")
            else:
                st.dataframe(
                    estilizar_tabela(
                        resumo,
                        ["SICC", "Primavera", "Diferença SICC - Primavera"],
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

                total_divergencias = int((resumo["Estado"] == "Divergência").sum())
                maior_diferenca = float(
                    resumo["Diferença SICC - Primavera"].abs().max()
                )
                c1, c2, c3 = st.columns(3)
                c1.metric("Componentes divergentes", total_divergencias)
                c2.metric("Maior diferença", formatar_euro(maior_diferenca))
                c3.metric("Bens a verificar", int(len(problemas)))

            st.subheader("Controlo das fichas")
            st.dataframe(resumo_controlo, use_container_width=True, hide_index=True)

        with separador_2:
            st.subheader("Reconciliação detalhada por conta")
            tabela_detalhe = detalhe.copy()

            if apenas_divergencias and not tabela_detalhe.empty:
                tabela_detalhe = tabela_detalhe[
                    tabela_detalhe["Estado"] == "Divergência"
                ]

            if tabela_detalhe.empty:
                st.success("Não existem divergências para os critérios selecionados.")
            else:
                tabela_detalhe = tabela_detalhe.sort_values(
                    "Diferença SICC - Primavera",
                    key=lambda s: s.abs(),
                    ascending=False,
                )
                st.dataframe(
                    estilizar_tabela(
                        tabela_detalhe,
                        ["SICC", "Primavera", "Diferença SICC - Primavera"],
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

        with separador_3:
            st.subheader("Possíveis bens sem depreciação ou amortização")
            st.caption(
                "São excluídos os terrenos. Nos restantes bens, se o valor de aquisição menos as amortizações/depreciações acumuladas for positivo, o valor do período tem de ser superior a zero."
            )

            if problemas.empty:
                st.success(
                    "Não foram identificados bens elegíveis sem depreciação/amortização ou com incoerências."
                )
            else:
                st.dataframe(
                    estilizar_tabela(
                        problemas,
                        [
                            "Valor de aquisição",
                            "Valor de aquisição - acumulada",
                            "Período",
                            "Exercício",
                            "Acumulada",
                        ],
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

        with separador_4:
            st.subheader("Cobertura do ficheiro SICC")
            st.dataframe(cobertura, use_container_width=True, hide_index=True)

            st.subheader("Lógica aplicada")
            st.markdown(
                """
- **AFT — valor contabilístico:** contas `431…437` do SICC, calculadas por **saldo a débito − saldo a crédito**, comparadas com o **Valor Contabilístico** do Primavera.
- **Ativos intangíveis — valor contabilístico:** contas `443…` do SICC, calculadas por **saldo a débito − saldo a crédito**, comparadas com o **Valor Contabilístico** do Primavera.
- **AFT — depreciação do período:** contas `642…`, calculadas por **valor a débito − valor a crédito**, usando primeiro o mapa contabilístico específico e só depois a inferência por raiz. Exemplo: `64259 → 4352 + 4353 + 4359`.
- **Ativos intangíveis — amortização do período:** contas `643…`, calculadas por **valor a débito − valor a crédito**, comparadas com a coluna **Período** do Primavera.
- **AFT — depreciação do exercício:** contas `642…`, calculadas por **saldo a débito − saldo a crédito**, comparadas com a soma das contas finais do Primavera pertencentes à respetiva raiz.
- **Ativos intangíveis — amortização do exercício:** contas `643…`, calculadas por **saldo a débito − saldo a crédito**, comparadas com a coluna **Exercício** do Primavera.
- **AFT — depreciação acumulada:** contas `438…`, calculadas por **saldo a crédito − saldo a débito**, usando o mapa específico quando várias contas de aquisição acumulam numa só conta. Exemplo: `43859 → 4352 + 4353 + 4359`.
- **Ativos intangíveis — amortização acumulada:** contas `4483…`, calculadas por **saldo a crédito − saldo a débito**, comparadas com a coluna **Acumulada** do Primavera.
- Nas contas hierárquicas são usadas apenas as contas finais do SICC, evitando a duplicação de contas-mãe e subcontas.
- **Bens sem amortização/depreciação:** os terrenos são excluídos. Para cada outro bem, calcula-se **Valor de aquisição − Acumulada**. Se o resultado for positivo, o valor da coluna **Período** tem de ser estritamente superior a zero; caso contrário, a ficha é assinalada como erro.
                """
            )

        relatorio_excel = gerar_excel(
            resumo,
            detalhe,
            cobertura,
            resumo_controlo,
            problemas,
            sicc,
            contas_primavera,
            fichas_primavera,
        )

        st.download_button(
            "Descarregar relatório Excel",
            data=relatorio_excel,
            file_name="relatorio_conferencia_ativos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as exc:
        st.error(f"Não foi possível processar os ficheiros: {exc}")
        st.exception(exc)

else:
    st.info("Carregue o balancete SICC e o balancete Primavera para iniciar a análise.")

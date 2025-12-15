import os
import re
from datetime import date
from typing import List, Dict, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment
from io import BytesIO


# =====================================================
# 1. Caminho do CSV de mapeamento Empresa → Entidade
# =====================================================

def get_mapping_path(default_path: str) -> str:
    """Tenta encontrar o ficheiro de mapeamento."""
    if os.path.isfile(default_path):
        return default_path
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidate = os.path.join(script_dir, "mapeamento_entidades_nc.csv")
    if os.path.isfile(candidate):
        return candidate
    return default_path

MAPPING_CSV_PATH = get_mapping_path("mapeamento_entidades_nc.csv")
ENTIDADE_PADRAO = "999"


# =====================================================
# 2. Cabeçalhos EXACTOS do ficheiro de importação
# =====================================================

HEADER = [
    "NC",
    "Entidade",
    "Data documento",
    "Data Contabilistica",
    "Nº NC",
    "Série",
    "Subtipo",
    "classificador economico ",
    "Classificador funcional ",
    "Fonte de financiamento ",
    "Programa ",
    "Medida",
    "Projeto",
    "Regionalização",
    "Atividade",
    "Natureza",
    "Departamento/Atividade",
    "Conta Debito",
    "Conta a Credito ",
    "Valor Lançamento",
    "Centro de custo",
    "Observações Documento ",
    "Observaçoes lançamento",
    "Classificação Orgânica",
    "Litigio",
    "Data Litigio",
    "Data Fim Litigio",
    "Plano Pagamento",
    "Data Plano Pagamento",
    "Data Fim Plano Pag",
    "Pag Factoring",
    "Nº Compromisso Assumido",
    "Projeto Documento",
    "Ano Compromisso Assumido",
    "Série Compromisso Assumido",
]


# =====================================================
# 3. Funções de base
# =====================================================

def load_empresa_mapping(path: str = MAPPING_CSV_PATH) -> pd.DataFrame:
    """Lê o CSV de mapeamento Empresa;Entidade."""
    df = pd.read_csv(path, sep=";", encoding="latin-1")
    df.columns = [str(c).strip() for c in df.columns]
    obrig = ["Empresa", "Entidade"]
    for c in obrig:
        if c not in df.columns:
            raise ValueError(f"O ficheiro de mapeamento tem de ter as colunas: {obrig}")
    return df


def normalizar_texto(s: str) -> str:
    """Normaliza texto para comparação."""
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    return s.upper()


def limpar_nome_coluna(col: str) -> str:
    """Remove HTML e caracteres especiais de nomes de colunas."""
    col = re.sub(r'<[^>]+>', '', str(col))
    col = col.strip()
    return col


def obter_mapa_empresas(mapping_df: pd.DataFrame) -> Dict[str, str]:
    """Cria dicionário de mapeamento Empresa → Entidade."""
    map_dict: Dict[str, str] = {}
    for _, row in mapping_df.iterrows():
        emp = normalizar_texto(row["Empresa"])
        ent = str(row["Entidade"]).strip()
        if not emp or not ent or ent.lower() in ("nan", "none"):
            continue
        map_dict[emp] = ent
    return map_dict


def separar_por_entidade(
    df_nc: pd.DataFrame, 
    mapping_df: pd.DataFrame
) -> Dict[str, Tuple[pd.DataFrame, List[str]]]:
    """Separa o DataFrame por entidade."""
    map_dict = obter_mapa_empresas(mapping_df)
    
    def get_entidade_para_linha(empresa):
        emp_norm = normalizar_texto(empresa)
        return map_dict.get(emp_norm, ENTIDADE_PADRAO)
    
    df_nc = df_nc.copy()
    df_nc['_entidade_calculada'] = df_nc['Empresa'].apply(get_entidade_para_linha)
    
    resultado = {}
    empresas_sem_mapa = []
    
    for entidade in df_nc['_entidade_calculada'].unique():
        df_ent = df_nc[df_nc['_entidade_calculada'] == entidade].copy()
        empresas_desta_ent = sorted(
            df_ent['Empresa'].dropna().map(normalizar_texto).unique()
        )
        
        for emp in empresas_desta_ent:
            if emp not in map_dict:
                empresas_sem_mapa.append(emp)
        
        resultado[entidade] = (df_ent.drop(columns=['_entidade_calculada']), empresas_desta_ent)
    
    resultado['_empresas_sem_mapa'] = empresas_sem_mapa
    return resultado


def detectar_formato_ficheiro(df: pd.DataFrame) -> Dict[str, str]:
    """Deteta formato do ficheiro e mapeia colunas."""
    colunas_limpas = {col: limpar_nome_coluna(col) for col in df.columns}
    colunas_disponiveis = list(colunas_limpas.values())
    
    mapeamento_alternativas = {
        "Data": ["Data", "Data Documento", "Data NC"],
        "Empresa": ["Empresa", "Nome Empresa"],
        "Instituição": ["Instituição", "Instituicao", "Cliente"],
        "Tipo": ["Tipo", "Tipo Documento"],
        "N.º / Ref.ª": ["N.º / Ref.ª", "Nº Documento", "Número", "Referência"],
        "Valor (com IVA)": ["Valor (com IVA)", "Valor", "Total"],
    }
    
    colunas_encontradas = {}
    colunas_faltantes = []
    
    for col_padrao, alternativas in mapeamento_alternativas.items():
        encontrada = False
        for alt in alternativas:
            if alt in colunas_disponiveis:
                for col_orig, col_limpa in colunas_limpas.items():
                    if col_limpa == alt:
                        colunas_encontradas[col_padrao] = col_orig
                        encontrada = True
                        break
                if encontrada:
                    break
        if not encontrada:
            colunas_faltantes.append(col_padrao)
    
    if colunas_faltantes:
        raise ValueError(
            f"Colunas obrigatórias não encontradas: {', '.join(colunas_faltantes)}.\n"
            f"Colunas disponíveis: {', '.join(colunas_disponiveis)}"
        )
    
    for col_orig, col_limpa in colunas_limpas.items():
        col_upper = col_limpa.upper()
        if "ANO" in col_upper and "Ano" not in colunas_encontradas:
            colunas_encontradas["Ano"] = col_orig
        if "TRANCHE" in col_upper and "Tranche" not in colunas_encontradas:
            colunas_encontradas["Tranche"] = col_orig
    
    return colunas_encontradas


def ler_notas_credito(file) -> pd.DataFrame:
    """Lê ficheiros de Notas de Crédito."""
    from io import StringIO
    
    fname = file.name.lower()

    if fname.endswith((".xlsx", ".xls")):
        file.seek(0)
        df = pd.read_excel(file)
    else:
        file.seek(0)
        raw = file.read()
        text = None
        
        for enc in ["utf-16", "utf-16-le", "utf-16-be", "utf-8", "utf-8-sig", "latin-1", "cp1252", "iso-8859-1"]:
            try:
                text = raw.decode(enc)
                break
            except (UnicodeDecodeError, AttributeError):
                continue

        if text is None:
            raise ValueError(f"Não foi possível decodificar '{file.name}'.")

        if text.startswith('\ufeff'):
            text = text[1:]

        lines = text.splitlines()
        if not lines:
            raise ValueError("Ficheiro vazio.")
            
        first_line = lines[0]
        skip = 1 if first_line.lower().strip().startswith(('sep=', '"sep=')) else 0
        sep = ";" if ";" in first_line else ("," if "," in first_line else "\t")

        df = pd.read_csv(StringIO(text), sep=sep, skiprows=skip)

    df = df.dropna(axis=1, how='all')
    mapeamento_colunas = detectar_formato_ficheiro(df)
    rename_dict = {v: k for k, v in mapeamento_colunas.items()}
    df = df.rename(columns=rename_dict)
    
    formato_info = []
    if "Ano" in mapeamento_colunas:
        formato_info.append("Ano")
    if "Tranche" in mapeamento_colunas:
        formato_info.append("Tranche")
    
    df.attrs['formato_detectado'] = ", ".join(formato_info) if formato_info else "formato básico"

    df_nc = df[df["Tipo"].astype(str).str.upper().str.contains("NOTA DE CRÉDITO", na=False)].copy()
    if df_nc.empty:
        raise ValueError("Nenhuma 'NOTA DE CRÉDITO' encontrada.")

    def parse_valor(v):
        s = str(v).strip()
        if not s or s.lower() in ("nan", "none"):
            return 0.0
        s = s.replace(" ", "")
        return float(s.replace(".", "").replace(",", "."))

    df_nc["ValorNum"] = df_nc["Valor (com IVA)"].apply(parse_valor)
    df_nc.attrs = df.attrs
    return df_nc


def format_yyyymmdd(data_str: str) -> str:
    """Converte data para AAAAMMDD."""
    s = str(data_str).strip()
    if "-" in s:
        partes = s.split("-")
        if len(partes) == 3:
            return partes[0] + partes[1] + partes[2]
    if "/" in s:
        partes = s.split("/")
        if len(partes) == 3:
            dia, mes, ano = partes
            return ano + mes.zfill(2) + dia.zfill(2)
    return s


def today_yyyymmdd() -> str:
    """Data de hoje em AAAAMMDD."""
    hoje = date.today()
    return f"{hoje.year:04d}{hoje.month:02d}{hoje.day:02d}"


def format_valor_port(valor: float) -> str:
    """1234.5 → '1234,50'"""
    return f"{valor:.2f}".replace(".", ",")


def apenas_algarismos(texto: str) -> str:
    """Apenas dígitos."""
    return re.sub(r"\D", "", str(texto))


def gerar_linhas_importacao_para_ficheiro(
    df_nc: pd.DataFrame,
    entidade: str,
    tipo_nc_prefix: str,
) -> List[List[str]]:
    """Gera linhas do CSV de importação."""
    linhas: List[List[str]] = []

    tem_ano = "Ano" in df_nc.columns
    tem_tranche = "Tranche" in df_nc.columns
    data_contab = today_yyyymmdd()

    for _, row in df_nc.iterrows():
        data_doc = format_yyyymmdd(row["Data"])
        valor = float(row["ValorNum"])
        ref = str(row["N.º / Ref.ª"])
        numero_nc = apenas_algarismos(ref)

        obs_parts = []
        if tem_ano and pd.notna(row["Ano"]):
            ano_val = str(row["Ano"]).strip()
            if ano_val and ano_val.lower() not in ("nan", "none", ""):
                obs_parts.append(ano_val)
        if tem_tranche and pd.notna(row["Tranche"]):
            tranche_val = str(row["Tranche"]).strip()
            if tranche_val and tranche_val.lower() not in ("nan", "none", ""):
                obs_parts.append(tranche_val)
        
        observacoes_base = " ".join(obs_parts).strip()
        observacoes_doc = f"{tipo_nc_prefix} {observacoes_base}".strip() if observacoes_base else tipo_nc_prefix

        linha: Dict[str, str] = {col: "" for col in HEADER}

        linha["NC"] = "NC"
        linha["Entidade"] = entidade
        linha["Data documento"] = data_doc
        linha["Data Contabilistica"] = data_contab
        linha["Nº NC"] = numero_nc
        linha["Série"] = ""
        linha["Subtipo"] = ""
        linha["classificador economico "] = "02.01.09.C0.00"
        linha["Classificador funcional "] = "0730"
        linha["Fonte de financiamento "] = "511"
        linha["Programa "] = "011"
        linha["Medida"] = "022"
        linha["Projeto"] = ""
        linha["Regionalização"] = ""
        linha["Atividade"] = "130"
        linha["Natureza"] = ""
        linha["Departamento/Atividade"] = "1"
        linha["Conta Debito"] = "221111"
        linha["Conta a Credito "] = "31826111"
        linha["Valor Lançamento"] = format_valor_port(valor)
        linha["Centro de custo"] = ""
        linha["Observações Documento "] = observacoes_doc
        linha["Observaçoes lançamento"] = ""
        linha["Classificação Orgânica"] = "101904000"
        linha["Litigio"] = ""
        linha["Data Litigio"] = ""
        linha["Data Fim Litigio"] = ""
        linha["Plano Pagamento"] = ""
        linha["Data Plano Pagamento"] = ""
        linha["Data Fim Plano Pag"] = ""
        linha["Pag Factoring"] = ""
        linha["Nº Compromisso Assumido"] = ""
        linha["Projeto Documento"] = ""
        linha["Ano Compromisso Assumido"] = ""
        linha["Série Compromisso Assumido"] = ""

        linhas.append([linha[col] for col in HEADER])

    return linhas


def escrever_excel_bytes(linhas: List[List[str]]) -> bytes:
    """
    Escreve ficheiro Excel com colunas formatadas como texto.
    ⚠️ Colunas 0730, 011, 022 são formatadas como TEXTO para preservar zeros.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Importação"
    
    # Identificar índices das colunas que devem ser texto
    indices_texto = set()
    colunas_texto = {
        "Classificador funcional ",  # 0730
        "Programa ",                  # 011
        "Medida"                      # 022
    }
    
    for i, col in enumerate(HEADER):
        if col in colunas_texto:
            indices_texto.add(i)
    
    # Escrever header
    ws.append(HEADER)
    
    # Escrever dados
    for linha in linhas:
        row_data = []
        for i, valor in enumerate(linha):
            if i in indices_texto and valor:
                # Forçar como texto adicionando apóstrofo invisível
                row_data.append(f"'{valor}")
            else:
                row_data.append(valor)
        ws.append(row_data)
    
    # Formatar colunas como texto
    for col_idx in indices_texto:
        col_letter = chr(65 + col_idx)  # A=65, B=66, etc
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            cell.number_format = '@'  # Formato texto
    
    # Salvar em memória
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# =====================================================
# 4. Interface Streamlit
# =====================================================

st.set_page_config(page_title="NC APIFARMA / PAYBACK → Importação", layout="wide")

st.title("Conversor de Notas de Crédito APIFARMA / PAYBACK")

st.markdown("""
Converte ficheiros de **Notas de Crédito** (Excel ou CSV) para importação contabilística.

**✨ Gera ficheiros Excel (.xlsx) com formatação correta de zeros à esquerda.**
""")

try:
    mapping_df = load_empresa_mapping(MAPPING_CSV_PATH)
    st.success(f"✅ Mapa carregado: {len(mapping_df)} empresas")
    with st.expander("Ver mapeamento"):
        st.dataframe(mapping_df, use_container_width=True)
except Exception as e:
    st.error(f"❌ Erro no mapeamento: {e}")
    st.stop()

st.divider()

st.header("1️⃣ Tipo de Nota de Crédito")
st.info("💡 Podes carregar ficheiros APIFARMA e PAYBACK ao mesmo tempo.")

col1, col2 = st.columns(2)
with col1:
    st.subheader("APIFARMA")
    uploaded_apifarma = st.file_uploader(
        "Ficheiros APIFARMA",
        type=["xlsx", "xls", "csv", "txt"],
        accept_multiple_files=True,
        key="apifarma_uploader"
    )

with col2:
    st.subheader("PAYBACK")
    uploaded_payback = st.file_uploader(
        "Ficheiros PAYBACK",
        type=["xlsx", "xls", "csv", "txt"],
        accept_multiple_files=True,
        key="payback_uploader"
    )

ficheiros_para_processar = []
if uploaded_apifarma:
    for f in uploaded_apifarma:
        ficheiros_para_processar.append((f, "APIFARMA"))
if uploaded_payback:
    for f in uploaded_payback:
        ficheiros_para_processar.append((f, "PAYBACK"))

if ficheiros_para_processar:
    st.header("2️⃣ Pré-visualização")
    preview_rows = []
    for file, tipo in ficheiros_para_processar:
        try:
            df_nc = ler_notas_credito(file)
            entidades_dict = separar_por_entidade(df_nc, mapping_df)
            empresas_sem_mapa = entidades_dict.pop('_empresas_sem_mapa', [])
            
            entidades_str = ", ".join(sorted(entidades_dict.keys()))
            formato = df_nc.attrs.get('formato_detectado', 'N/A')
            
            status = "✅ OK"
            if empresas_sem_mapa:
                status += f" (⚠️ {len(empresas_sem_mapa)} → 999)"
            
            preview_rows.append({
                "Ficheiro": file.name,
                "Tipo": tipo,
                "Entidades": entidades_str,
                "Formato": formato,
                "NCs": len(df_nc),
                "Estado": status
            })
        except Exception as e:
            preview_rows.append({
                "Ficheiro": file.name,
                "Tipo": tipo,
                "Entidades": "",
                "Formato": "",
                "NCs": 0,
                "Estado": f"❌ {str(e)[:50]}..."
            })

    st.dataframe(pd.DataFrame(preview_rows), use_container_width=True)

process_button = st.button("▶️ Converter ficheiros", type="primary")

if process_button:
    if not ficheiros_para_processar:
        st.error("❌ Carrega pelo menos um ficheiro.")
    else:
        st.header("3️⃣ Ficheiros gerados")
        
        todas_empresas_sem_mapa = set()
        
        for file, tipo_nc_prefix in ficheiros_para_processar:
            st.subheader(f"📄 {file.name} ({tipo_nc_prefix})")
            
            try:
                df_nc = ler_notas_credito(file)
                entidades_dict = separar_por_entidade(df_nc, mapping_df)
                empresas_sem_mapa = entidades_dict.pop('_empresas_sem_mapa', [])
                
                if empresas_sem_mapa:
                    todas_empresas_sem_mapa.update(empresas_sem_mapa)
                    st.warning(
                        f"⚠️ {len(empresas_sem_mapa)} empresa(s) → código {ENTIDADE_PADRAO}: "
                        f"{', '.join(empresas_sem_mapa)}"
                    )
                
                todas_linhas_ficheiro = []
                total_notas_ficheiro = 0
                
                for entidade, (df_ent, empresas) in entidades_dict.items():
                    total_notas_ficheiro += len(df_ent)
                    linhas = gerar_linhas_importacao_para_ficheiro(df_ent, entidade, tipo_nc_prefix)
                    todas_linhas_ficheiro.extend(linhas)
                
                formato = df_nc.attrs.get('formato_detectado', 'desconhecido')
                
                st.success(
                    f"✅ **Processado!**\n\n"
                    f"- NCs: {total_notas_ficheiro}\n"
                    f"- Linhas: {len(todas_linhas_ficheiro)}\n"
                    f"- Formato: {formato}"
                )
                
                nome_base = os.path.splitext(file.name)[0]
                nome_saida = f"NC_{tipo_nc_prefix}_{nome_base}_importacao.xlsx"
                
                excel_bytes = escrever_excel_bytes(todas_linhas_ficheiro)
                
                st.download_button(
                    f"⬇️ Descarregar {nome_saida}",
                    excel_bytes,
                    nome_saida,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_{file.name}"
                )
                
                st.info("💡 **Nota:** Abre o ficheiro Excel, verifica os dados e depois grava como CSV se necessário.")
                
            except Exception as e:
                st.error(f"❌ Erro: {e}")
        
        if todas_empresas_sem_mapa:
            st.divider()
            st.warning(
                f"💡 **{len(todas_empresas_sem_mapa)} empresa(s) sem mapeamento.**\n\n"
                f"Empresas: {', '.join(sorted(todas_empresas_sem_mapa))}"
            )
